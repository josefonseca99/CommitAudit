import os
import xlsxwriter
import openpyxl
import pandas
from openpyxl.styles import Alignment
from util import ConstantManagement, GetAnalystsTasks, ExtractPath
from openpyxl.styles.borders import Border, Side
from openpyxl import styles


def end_of_file_position():
    excel_file = openpyxl.load_workbook(ExtractPath.relative_path(ConstantManagement.EXCEL_FILE_NAME))
    sheet_name = excel_file.sheetnames[0]
    worksheet = excel_file[sheet_name]
    excel_file.close()
    return worksheet.max_row


def audit_file_exists():
    return os.path.exists(ExtractPath.relative_path(ConstantManagement.EXCEL_FILE_NAME))


class CreateExcel:

    def __init__(self, sprint_number=None):
        self.sprint_number = sprint_number
        if not audit_file_exists():
            self.book = xlsxwriter.Workbook(ConstantManagement.EXCEL_FILE_NAME)
            self.sheet = self.book.add_worksheet()
            self.header_format = self.book.add_format({'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#0B3861'})
            self.header_format.set_border()
            self.header_format.set_align('center')
            self.basic_format = self.book.add_format({'bold': False, 'font_color': '#000000', 'bg_color': '#FFFFFF'})
            self.basic_format.set_border()
            self.basic_format.set_align('center')
            self.header_format.set_align('vcenter')
            self.header_format.set_text_wrap()
            self.excel_header()
            self.submit_result_values()
            self.book.close()
        else:
            self.update_audit_file()

    def excel_header(self):

        self.sheet.set_column(0, len(ConstantManagement.HEADER_VALUES) - 1, width=40)
        self.sheet.set_row(0, height=60)

        for header in range(0, len(ConstantManagement.HEADER_VALUES)):
            self.sheet.write(0, header, ConstantManagement.HEADER_VALUES[header], self.header_format)

    def submit_result_values(self):
        task_list_creation = GetAnalystsTasks.GetAnalystsTasks(self.sprint_number)
        task_list_creation.collect_data()
        analysts_tasks_list = task_list_creation.get_data_collection()

        for row in range(0, len(analysts_tasks_list)):
            self.sheet.write(row + 1, 0, self.sprint_number, self.basic_format)
            self.sheet.write(row + 1, 1, analysts_tasks_list[row]["analyst_name"], self.basic_format)
            self.sheet.write(row + 1, 2, analysts_tasks_list[row]["enabler_type"], self.basic_format)
            self.sheet.write(row + 1, 3, analysts_tasks_list[row]["user_story_type"], self.basic_format)
            self.sheet.write(row + 1, 4, analysts_tasks_list[row]["bug_type"], self.basic_format)
            self.sheet.write(row + 1, 5, analysts_tasks_list[row]["new_state"], self.basic_format)
            self.sheet.write(row + 1, 6, analysts_tasks_list[row]["active_state"], self.basic_format)
            self.sheet.write(row + 1, 7, analysts_tasks_list[row]["closed_state"], self.basic_format)
            self.sheet.write(row + 1, 8, analysts_tasks_list[row]["impediment_state"], self.basic_format)
            self.sheet.write(row + 1, 9, analysts_tasks_list[row]["engaged_to"], self.basic_format)
            self.sheet.write(row + 1, 10, str(analysts_tasks_list[row]["no_scored"]).strip("[]"), self.basic_format)
            self.sheet.write(row + 1, 11, analysts_tasks_list[row]["pull_data"], self.basic_format)
            self.sheet.write(row + 1, 12, analysts_tasks_list[row]["commit_data"], self.basic_format)

    def update_audit_file(self):
        task_list_creation = GetAnalystsTasks.GetAnalystsTasks(self.sprint_number)
        task_list_creation.collect_data()
        analysts_tasks_list = task_list_creation.get_data_collection()

        excel_file = openpyxl.load_workbook(ExtractPath.relative_path(ConstantManagement.EXCEL_FILE_NAME))
        sheet_name = excel_file.sheetnames[0]
        worksheet = excel_file[sheet_name]

        eof_position = end_of_file_position()

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        align_center = Alignment(horizontal='center')

        for file_row in range(0, len(analysts_tasks_list)):
            row_position = file_row + eof_position + 1

            worksheet.cell(row=row_position, column=1, value=self.sprint_number)
            worksheet.cell(row=row_position, column=2, value=analysts_tasks_list[file_row]["analyst_name"])
            worksheet.cell(row=row_position, column=3, value=analysts_tasks_list[file_row]["enabler_type"])
            worksheet.cell(row=row_position, column=4, value=analysts_tasks_list[file_row]["user_story_type"])
            worksheet.cell(row=row_position, column=5, value=analysts_tasks_list[file_row]["bug_type"])
            worksheet.cell(row=row_position, column=6, value=analysts_tasks_list[file_row]["new_state"])
            worksheet.cell(row=row_position, column=7, value=analysts_tasks_list[file_row]["active_state"])
            worksheet.cell(row=row_position, column=8, value=analysts_tasks_list[file_row]["closed_state"])
            worksheet.cell(row=row_position, column=9, value=analysts_tasks_list[file_row]["impediment_state"])
            worksheet.cell(row=row_position, column=10, value=analysts_tasks_list[file_row]["engaged_to"])
            worksheet.cell(row=row_position, column=11,
                           value=str(analysts_tasks_list[file_row]["no_scored"]).strip("[]"))
            worksheet.cell(row=row_position, column=12, value=analysts_tasks_list[file_row]["pull_data"])
            worksheet.cell(row=row_position, column=13, value=analysts_tasks_list[file_row]["commit_data"])

            for file_column in range(1, len(ConstantManagement.HEADER_VALUES)+1):
                worksheet.cell(row=row_position, column=file_column).border = thin_border
                worksheet.cell(row=row_position, column=file_column).alignment = align_center

        excel_file.save(ExtractPath.relative_path(ConstantManagement.EXCEL_FILE_NAME))
        excel_file.close()
