import openpyxl
from util import ConstantManagement, ExtractPath


class GetExcelAnalysts:
    def __init__(self):
        self.provider_data = openpyxl.load_workbook(ExtractPath.relative_path(ConstantManagement.EXCEL_ANALYSTS))
        self.sheet_name = self.provider_data.sheetnames[0]
        self.worksheet = self.provider_data[self.sheet_name]
        self.analyst_number = self.worksheet['A']
        self.analysts_list = []
        for index in range(len(self.analyst_number) - 1):
            cell_index = index + 2
            self.analysts_list.append(
                {"name": self.worksheet.cell(cell_index, 1).value, "email": self.worksheet.cell(cell_index, 2).value, "cell": self.worksheet.cell(cell_index, 3).value,
                 "repository": self.worksheet.cell(cell_index, 4).value})

    def analysts_dict(self):
        return self.analysts_list
